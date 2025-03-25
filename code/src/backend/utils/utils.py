from langchain_google_genai import ChatGoogleGenerativeAI
from langchain_ollama import OllamaLLM
import pandas as pd
from langchain_community.document_loaders import PyPDFLoader
from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain_core.vectorstores import InMemoryVectorStore
from langchain.prompts import PromptTemplate
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment
import io
import re

def load_llm(model_type,model_name,api_key=None):
    if model_type.lower()=="google":
        return ChatGoogleGenerativeAI(model=model_name,api_key=api_key)
    elif model_type.lower()=="ollama":
        return OllamaLLM(model=model_name)
    else:
        raise ValueError("Unsupported model type. Choose 'google' or 'ollama'.")
        
def load_data(file_path):
    return pd.read_csv(file_path)
    
def get_column_names(data):
    return data.columns.tolist()

def extract_data_chunks(pdf_path):
    #load the pdf file
    loader = PyPDFLoader(pdf_path)
    documents = loader.load()
    
    # Define chunking parameters
    text_splitter = RecursiveCharacterTextSplitter(
    chunk_size=4000,  # Size of each chunk
    chunk_overlap=400  # Overlap between chunks to preserve context
    )
    return text_splitter.split_documents(documents=documents)

def load_vector_store(embedding_model):
    return InMemoryVectorStore(embedding_model)
    
def get_document_embeddings(vector_store, document_chunks):
    _ = vector_store.add_documents(documents=document_chunks)
    return vector_store
    
def prompt_template(type):
    if(type=="data_profiling_rules"):
        return PromptTemplate(
    input_variables=["column"],
    template="""
    For the columns '{column}', generate a data profiling rule that ensures data quality and consistency based on the regulations found in {context}.
    Ensure the rule accounts for business exceptions where applicable.
    
    Example rules:
    1) Transaction_Amount should always match Reported_Amount, except when the transaction involves cross-currency conversions, in which case a permissible deviation of up to 1% is allowed.
    2) Account_Balance should never be negative, except in cases of overdraft accounts explicitly marked with an "OD" flag.
    3) Currency should be a valid ISO 4217 currency code, and the transaction must adhere to cross-border transaction limits as per regulatory guidelines.
    4) Country should be an accepted jurisdiction based on bank regulations, and cross-border transactions should include mandatory transaction remarks if the amount exceeds $10,000.

    Now generate profiling  rules only for the columns mentioned in'{column}'  based on  {context}. Number the columns sequentially. Give as a plain text without any formatting.  
    Output Format : 
    Here are the data profiling rules for the columns :
    Column Name : Rules for the column
    
    Example :
    
    1) Transaction_Amount - should always match Reported_Amount, except when the transaction involves cross-currency conversions, in which case a permissible deviation of up to 1% is allowed.
    2) Account_Balance - should never be negative, except in cases of overdraft accounts explicitly marked with an "OD" flag.
    
    """
    )

def process_llm_output(raw_data):
    # Split based on numbered sections
    sections = re.split(r'(\d+\))', raw_data)
    # Process each section
    formatted_output = []
    for i in range(1, len(sections), 2):
        header = sections[i].strip()
        content = sections[i+1].strip().replace('*', '-').replace('\n', '\n    ')
        formatted_output.append(f"{header} {content}")
    # Join everything into a clean output
    return '\n\n'.join(formatted_output)

# def parse_violations(text):
#     """
#     Parses a block of text containing violations, remediations, and metadata into a structured dictionary.
    
#     :param text: The raw text input with violations, remediations, and metadata.
#     :return: A dictionary with structured violations, remediations, and metadata.
#     """
#     parsed_data = {"violations": {}, "remediations": {}}

#     # Extract all violations and remediations
#     violation_matches = re.findall(r"Violation: (.+?): (.+?)(?=\nViolation:|$)", text)
#     remediation_matches = re.findall(r"Remediation: (.+?): (.+?)(?=\nRemediation:|$)", text)

#     # Populate violations dictionary
#     for field, message in violation_matches:
#         parsed_data["violations"][field.strip()] = message.strip()

#     # Populate remediations dictionary
#     for field, message in remediation_matches:
#         parsed_data["remediations"][field.strip()] = message.strip()

#     # Extract boolean and numeric values
#     for line in text.split("\n"):
#         if "violation_exists" in line:
#             parsed_data["violation_exists"] = line.split(": ")[1].strip() == "True"
#         elif "risk_score" in line:
#             parsed_data["risk_score"] = int(line.split(": ")[1].strip())

#     return parsed_data
def parse_violations(text):
    """
    Parses violations, remediations, and metadata into a structured dictionary.
    
    :param text: Raw input string with violations, remediations, and metadata.
    :return: Dictionary with violations, remediations, and metadata parsed.
    """
    parsed_data = {"violations": {}, "remediations": {}}

    # Extract violations and remediations
    violation_matches = re.findall(r"Violation: (.+?): (.+?)(?=\nViolation:|$)", text)
    remediation_matches = re.findall(r"Remediation: (.+?): (.+?)(?=\nRemediation:|$)", text)

    # Populate violations dictionary
    for field, message in violation_matches:
        parsed_data["violations"][field.strip()] = message.strip()

    # Populate remediations dictionary
    for field, message in remediation_matches:
        parsed_data["remediations"][field.strip()] = message.strip()

    # Extract boolean and numeric metadata
    for line in text.split("\n"):
        if "violation_exists" in line:
            parsed_data["violation_exists"] = line.split(": ")[1].strip().lower() == "true"
        elif "risk_score" in line:
            parsed_data["risk_score"] = int(line.split(": ")[1].strip())

    return parsed_data
# def parse_violations(text):
#     parsed_data = {"violations": {}, "remediations": {}}

#     # Extract all violations and remediations
#     violation_matches = re.findall(r"Violation: (.+?): (.+?)(?=\nRemediation:|$)", text)
#     remediation_matches = re.findall(r"Remediation: (.+?)(?=\nViolation:|$)", text)

#     # Populate the parsed_data dictionary
#     for violation in violation_matches:
#         field, message = violation
#         parsed_data["violations"][field.strip()] = message.strip()

#     # Match remediation steps with corresponding violations
#     for idx, remediation in enumerate(remediation_matches):
#         field = list(parsed_data["violations"].keys())[idx]
#         parsed_data["remediations"][field] = remediation.strip()

#     # Extract boolean and numeric values
#     for line in text.split("\n"):
#         if "violation_exists" in line:
#             parsed_data["violation_exists"] = line.split(": ")[1].strip() == "True"
#         elif "risk_score" in line:
#             parsed_data["risk_score"] = int(line.split(": ")[1].strip())

#     return parsed_data
    
# def store_violations(df, responses, folder):
#     print("Generating Excel file with highlighted violations and comments for all rows...")
#     df.to_excel(f"{folder}/violations_output.xlsx", index=False)
#     print("Excel file created with all rows!")
#     wb = load_workbook(f"{folder}/violations_output.xlsx")
#     ws = wb.active
#     red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

#     for response in responses:
#         row_idx = response['row_index']
#         print(response['violations_data'])
#         violation = response['violations_data']['violations_data'][0]['violation']
#         remediation = response['violations_data']['violations_data'][0]['remediation']
#         print(f"Row {row_idx + 1} has a violation: {violation}")
#         if response['violations_data']['violation_exists']:
#             # Color the entire row red
#             for col_num in range(1, len(df.columns) + 1):
#                 cell = ws.cell(row=row_idx + 1, column=col_num)
#                 cell.fill = red_fill
#                 print("cell filled")

#             # Add a comment with the single violation and remediation
#             comment_text = f"Violation: {violation}\nRemediation: {remediation}"
#             first_cell = ws.cell(row=row_idx + 1, column=1)
#             first_cell.comment = Comment(comment_text, "CheckerBot")

#     wb.save(f"{folder}/violations_output.xlsx")
#     print("Excel file generated with highlighted violations and comments for all rows!")
def store_violations(df, responses, folder):
    print("Generating Excel file with highlighted violations and comments for all rows...")
    df.to_excel(f"{folder}/violations_output.xlsx", index=False)
    print("Excel file created with all rows!")

    # Load the workbook and set up styling
    wb = load_workbook(f"{folder}/violations_output.xlsx")
    ws = wb.active
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    for response in responses:
        row_idx = response['row_index']
        violations_data = response['violations_data']

        # Check if the row has violations
        if violations_data['violation_exists']:
            print(f"Row {row_idx + 1} has violations!")               
            # Create a detailed comment for each violated field
            comments = []
            for field, violation_msg in violations_data['violations'].items():
                # check if remediation exists
                remediation_msg = violations_data['remediations'].get(field, "No remediation provided.")
                comments.append(f"{field}: {violation_msg}\nRemediation: {remediation_msg}")
                cell = ws.cell(row=row_idx + 1, column=df.columns.get_loc(field) + 1)
                cell.fill = red_fill
                cell.comment = Comment(f"Violation: {violation_msg}\nRemediation: {remediation_msg}", "CheckerBot")

    # Save the updated file
    wb.save(f"{folder}/violations_output.xlsx")
    print("Excel file generated with highlighted violations and detailed comments for all rows!")


    
    